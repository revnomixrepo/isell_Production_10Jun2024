# -*- coding: utf-8 -*-
"""
Created on Wed Jan 16 16:48:13 2019

@author: revse
"""
import os
import pandas as pd
import numpy as np
import logging
from openpyxl import Workbook
from datetime import datetime
from openpyxl.utils.dataframe import dataframe_to_rows
# from openpyxl.utils import coordinate_from_string, column_index_from_string
from openpyxl.utils import column_index_from_string
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.utils import get_column_letter
from datetime import date,timedelta

#ddmmyy=datetime.now()
#ddmmyy=date.today()+timedelta(days=-1)#A.S Jun23

to_day=date.today()
dow = to_day.strftime('%a')
tday=to_day.strftime('%d-%b-%Y')

iseldt = to_day.strftime('%d%b%Y')
foldname = to_day.strftime('%d_%b_%Y')

def beautify(defaultpath, df,iselltype,rowlim,htlname,pth,glossary,ftr,pgdf,finaladop,rateshopfile,hcap2,cm_room_type_df,tday,ls):
    logging.debug('------------------------------------------------------------')
    logging.debug('Module:beautiMode.py, SubModule:beautify')  
    os.chdir(pth)
    df = pd.DataFrame(df)
    try:
        df.drop(columns=np.nan, inplace=True)
    except KeyError:
        pass

    logging.debug('df contains isell dataframe ::')
    logging.debug(df)
    #=============================df to wb conv =============================================   
    logging.debug('dataframe to Workbook conversion:')
    wb = Workbook()    
    ws = wb.active
    logging.debug('Workbook(wb) and worksheet(ws) objects created')
    
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
        
    logging.debug('appended each row of df to worksheet object (ws)')
    #===================== column names and index and values===================================          
    col_index = [cell.column for cell in ws[1]]
    col_name = [cell.value for cell in ws[1]]
    
    logging.debug('index -> column_name  and column_name -> index dictionaries created ::')
    ind_name = dict(zip(col_index,col_name))
    logging.debug(ind_name)    
    
    name_ind = dict(zip(col_name,col_index))   
    logging.debug(name_ind)


    vals = []
    for i in range(1,len(col_index)):
        xy = get_column_letter(i)
        colvalue = column_index_from_string(xy[0])
        vals.append(colvalue)
    logging.debug('name_vals dictionary ::')
    name_vals = dict(zip(col_name, vals))

    #=============================================================================
    if iselltype == 'internal':
        writer = pd.ExcelWriter('iSell_{}_{}-Copy.xlsx'.format(htlname,iseldt), engine='xlsxwriter')
        logging.debug('Created writer object with iSell-Copy.xlsx')
    else:
        writer = pd.ExcelWriter('iSell_{}_{}.xlsx'.format(htlname,iseldt), engine='xlsxwriter')
        logging.debug('Created writer object with iSell.xlsx')
        
    #===============df positioning to 6th row================================================
    df.to_excel(writer, sheet_name='iSell_{}'.format(iseldt),startrow=6, startcol=0, header=False, index=False)
    logging.debug('pushed isell dataframe df to writer object at 6th row')


    # Get the xlsxwriter objects from the dataframe writer object.
    workbook  = writer.book
    worksheet = writer.sheets['iSell_{}'.format(iseldt)]
    logging.debug('first sheet name provided as :iSell_{}'.format(iseldt))


    #============================header format===========================================
    header_format = workbook.add_format({
            'bold':True,
            'text_wrap':True,
            'align':'center',
            'valign':'vcenter',
            'bg_color':	'#333333',
            'font_color': '#FFFFFF',
            'border':1       
            })
    
    #========================================================================================
    format1 = workbook.add_format({'bg_color': '#FF0000',
                                   'align':'right',
                                   'border':1,
                                   'font_color': '#000000'}) #Red
    
    # Add a format. Green fill with dark green text.
    format2 = workbook.add_format({'bg_color': '#00FF00',
                                   'align':'right',
                                   'border':1,
                                   'font_color': '#000000'}) #Green

    #===================================================
    format_bg12 = workbook.add_format({'bg_color': '#8DB4E2',
                                      'border': 1,
                                      'align': 'right',
                                      'font_color': '#000000'})  # pick blue
    format_bg21 = workbook.add_format({'bg_color': '#B4C69B',
                                      'border': 1,
                                      'align': 'right',
                                      'font_color': '#000000'})  # pick green
    #=====================================================
    
    format3 = workbook.add_format({'bg_color': '#FF8080',
                                   'border':1,
                                   'font_color': '#000000'}) #Pink
    
    format4 = workbook.add_format({'bg_color': '#FF99CC',
                                   'border':1,
                                   'align':'right',
                                   'font_color': '#000000'}) #darkPink
    
    format5 = workbook.add_format({'bg_color': '#FFCC99',
                                   'border':1,
                                   'align':'right',
                                   'font_color': '#000000'}) #thinOrange
        
    format6 = workbook.add_format({'bg_color': '#99CC00',
                                   'border':1,
                                   'align':'right',
                                   'font_color': '#000000'}) #thingreen
        
    format7 = workbook.add_format({'bg_color': '#99CCFF',
                                   'border':1,
                                   'align':'right',
                                   'font_color': '#000000'}) #thinblue
    
    format8 = workbook.add_format({'bg_color':'#FFFF00',
                                   'border':1,
                                   'align':'right',
                                   'font_color': '#000000'}) #yellow


    format9 = workbook.add_format({'bg_color':'#FFFF00',
                                   'border':1,
                                   'align':'right',
                                   'font_color': '#FFC7CE'}) #light red






    border_form = workbook.add_format({
                                   'border':1,
                                   'align':'right',
                                   })

    format11 = workbook.add_format({'bg_color': '#99CCFF',
                                    'bottom': 1, 'top': 1,
                                    'left': 0, 'right': 1,
                                    'align': 'right',
                                    'font_color': '#963634'})  # font Red left no border

    format12 = workbook.add_format({'bg_color': '#99CCFF',
                                    'bottom': 1, 'top': 1,
                                    'left': 0, 'right': 1,
                                    'align': 'right',
                                    'font_color': '#76933C'})  # font Green left no border

    format13 = workbook.add_format({'bg_color': '#99CCFF',
                                    'bottom': 1, 'top': 1,
                                    'left': 0, 'right': 1,
                                    'align': 'right',
                                    'font_color': '#ffffff'})  # font white left no border

    #==========================A.S. Jun23=================================
    format_date_1 = workbook.add_format({'bold': True,
                                         'border': 1,
                                         'align': 'right',
                                         'font_color': '#000000',
                                         'font_size': 10})

    format_date_2 = workbook.add_format({'bold': False,
                                         'border': 1,
                                         'align': 'right',
                                         'font_color': '#717171'}) # '#6e6d6d'

    format_date_3 = workbook.add_format({'bold': False,
                                         'border': 1,
                                         'align': 'right',
                                         'font_color': '#343434'})
    #===========================================================
    no_border_form = workbook.add_format({
        'border': 0,
        'align': 'right'})#for 0.5width columns

    header_format_blk_l = workbook.add_format({
        'bold': True,
        'text_wrap': True,
        'align': 'center',
        'valign': 'vcenter',
        'bg_color': '#333333',
        'font_color': '#333333',
        'bottom': 1,
        'top': 1,
        'left': 0,
        'right': 1
    })


    logging.debug('header,border and other formats are loaded')
    #==================== loop through 1st row (column names =========================================
    logging.debug('looping in iSell columns and assigning conditional formatting ...')
    mktrnd=df.columns.get_loc('Market Trend')
    lwrate=df.columns.get_loc('Lowest Rate')
    for names in col_name:
        colnum = name_ind[names]
        colname = ind_name[colnum]
        
        #========================== Event ============================================================
        if colname == 'Event':
            worksheet.conditional_format('{}7:{}{}'.format(colnum,colnum,rowlim),
                                     {'type': 'cell',
                                      'criteria' : '>',
                                      'value' : 0,
                                     'format': format3})

            worksheet.set_column('{}2:{}{}'.format(colnum,colnum,rowlim),30)
            
            worksheet.conditional_format('{}7:{}{}'.format(colnum,colnum,rowlim),
                                         {'type': 'cell',
                                          'criteria' : '<=',
                                          'value' : 0,
                                         'format': border_form})

            logging.debug('{} Column Conditional Formatting Done.'.format(colname))
        #===========================Capacity=================================================

        elif colname == 'Capacity':
            worksheet.conditional_format('{}7:{}{}'.format(colnum,colnum,rowlim),
                                     {'type': 'cell',
                                      'criteria' : '>',
                                      'value' : 0,
                                     'format': border_form})

    
        #========================= 'Dow' =====================================================
        elif colname == 'Dow':
            worksheet.conditional_format('{}7:{}{}'.format(colnum,colnum,rowlim),
                                         {'type': 'text',
                                          'criteria' : 'containing',
                                          'value' : 'Sat',
                                         'format': format4}) 
            
            worksheet.conditional_format('{}7:{}{}'.format(colnum,colnum,rowlim),
                                         {'type': 'text',
                                          'criteria' : 'containing',
                                          'value' : "Sun",
                                         'format': format4}) 
    
            worksheet.conditional_format('{}7:{}{}'.format(colnum,colnum,rowlim),
                                         {'type': 'cell',
                                          'criteria' : '>',
                                          'value' : 0,
                                         'format': border_form})
            logging.debug('{} Column Conditional Formatting Done.'.format(colname))
    
        #========================= 'Hotel Availability' =====================================================
        elif colname == 'Hotel Availability':
            worksheet.conditional_format('{}7:{}{}'.format(colnum,colnum,rowlim),
                                         {'type': 'cell',
                                          'criteria' : '<=',
                                          'value' : hcap2,
                                         'format': format8})  
        
            worksheet.conditional_format('{}7:{}{}'.format(colnum,colnum,rowlim),
                                         {'type': 'cell',
                                          'criteria' : '>',
                                          'value' : 0,
                                         'format': border_form})
    
            logging.debug('{} Column Conditional Formatting Done.'.format(colname))
            
        #========================= Rms Avail to Sell =====================================================
        elif colname == 'Rooms Avail To Sell Online':
            worksheet.conditional_format('{}7:{}{}'.format(colnum,colnum,rowlim),
                                         {'type': 'cell',
                                          'criteria' : '<=',
                                          'value' : 0,
                                         'format': format1})

            worksheet.conditional_format('{}7:{}{}'.format(colnum,colnum,rowlim),
                                         {'type': 'cell',
                                          'criteria' : '>',
                                          'value' : 0,
                                         'format': border_form})

            logging.debug('{} Column Conditional Formatting Done.'.format(colname))


            # ========================= Inventory Open% A.S.23 =====================================================
        elif colname == 'Inventory Open%':
            worksheet.conditional_format('{}7:{}{}'.format(colnum, colnum, rowlim),
                                         {'type': 'cell',
                                          'criteria': '<=',
                                          'value': 30,
                                          'format': format1})

            worksheet.conditional_format('{}7:{}{}'.format(colnum, colnum, rowlim),
                                         {'type': 'cell',
                                          'criteria': '>',
                                          'value': 30,
                                          'format': border_form})

            logging.debug('{} Column Conditional Formatting Done.'.format(colname))


         #========================= FTR name get from conditions =====================================================
        elif colname == ftr:
            worksheet.conditional_format('{}7:{}{}'.format(colnum,colnum,rowlim),
                                         {'type': 'cell',
                                          'criteria' : '<=',
                                          'value' : 0,
                                         'format': format1})

            worksheet.conditional_format('{}7:{}{}'.format(colnum,colnum,rowlim),
                                         {'type': 'cell',
                                          'criteria' : '>',
                                          'value' : 0,
                                         'format': border_form})

            worksheet.conditional_format('{}7:{}{}'.format(colnum,colnum,rowlim),
                                         {'type': 'cell',
                                          'criteria' : '<',
                                          'value' : 0,
                                         'format': border_form})

            logging.debug('{} Column Conditional Formatting Done.'.format(colname))
            
        #========================== OTA Sold (G) =============================================================
        elif 'Sold' in str(colname):
            worksheet.conditional_format('{}7:{}{}'.format(colnum,colnum,rowlim),
                                         {'type': 'cell',
                                          'criteria' : '>=',
                                          'value' : 0,                              
                                         'format': format4})
    
            logging.debug('{} Column Conditional Formatting Done.'.format(colname))
        #--------------------------------------------------------------------------------------------
        # elif colname=='Total Pickup':
        #     worksheet.conditional_format('{}7:{}{}'.format(colnum, colnum, rowlim),
        #                                  {'type': 'cell',
        #                                   'criteria': '<',
        #                                   'value': 0,
        #                                   'format': format1})
        #
        #     worksheet.conditional_format('{}7:{}{}'.format(colnum, colnum, rowlim),
        #                                  {'type': 'cell',
        #                                   'criteria': '>',
        #                                   'value': 0,
        #                                   'format': format2})
        #
        #     worksheet.conditional_format('{}7:{}{}'.format(colnum, colnum, rowlim),
        #                                  {'type': 'cell',
        #                                   'criteria': '=',
        #                                   'value': 0,
        #                                   'format': border_form})

            logging.debug('{} Column Conditional Formatting Done.'.format(colname))
        #=========================== PickupColumn(H) ===================================================================
        elif colname in ['Pickup','Total Pickup','Revenue Pickup','ADR Pickup','Rooms Last 7 Days Pickup','Revenue Last 7 Days Pickup','ADR Last 7 Days Pickup']:
            worksheet.conditional_format('{}7:{}{}'.format(colnum,colnum,rowlim),
                                         {'type': 'cell',
                                          'criteria' : '<',
                                          'value' : 0,
                                         'format': format_bg12})
            
            worksheet.conditional_format('{}7:{}{}'.format(colnum,colnum,rowlim),
                                         {'type': 'cell',
                                          'criteria' : '>',
                                          'value' : 0,
                                         'format': format_bg21})
        
            worksheet.conditional_format('{}7:{}{}'.format(colnum,colnum,rowlim),
                                         {'type': 'cell',
                                          'criteria' : '=',
                                          'value' : 0,
                                         'format': border_form})


        
        
            logging.debug('{} Column Conditional Formatting Done.'.format(colname))
        
         #=========================== ADR, Revenue, Rate on CM, ReccRate ===================================================================
       
        elif colname in ['Total Revenue','Total ADR OTB','OTA Revenue','ADR OTB','Rate on CM','Recommended Rate','Lowest Rate']:
            worksheet.conditional_format('{}7:{}{}'.format(colnum,colnum,rowlim),#rate on cm removed from list
                                         {'type': 'cell',
                                          'criteria' : '>=',
                                          'value' : 0,
                                         'format': format5})
            
            logging.debug('{} Column Conditional Formatting Done.'.format(colname))


        # elif colname in col_name[:name_vals['Market Trend']]:
        #     pass
        # elif (colname in (df.columns[lwrate:mktrnd])):
        # elif colname in ['Rate on CM']:
        #     worksheet.conditional_format('{}9:{}{}'.format(colnum, colnum, rowlim),
        #                                 {'type': 'formula',
        #                                  'criteria': '=${}9>={lwrate+1}'.format(colnum),
        #                                  'value': lwrate+1,
        #                                  'format': format5})
        #
        #     worksheet.conditional_format('{}9:{}{}'.format(colnum, colnum, rowlim),
        #                                 {'type': 'formula',
        #                                  'criteria': '=${}9<{lwrate+1}'.format(colnum),
        #                                  'value': lwrate+1,
        #                                  'format': border_form})
        #================================Date==A.S. Jun23============================================================================
        elif colname == 'Date':
            worksheet.conditional_format('{}7:{}{}'.format(colnum, colnum, rowlim),
                                         {'type': 'date',
                                          'criteria': '>',
                                          'value': tday,
                                          'format': format_date_3})

            worksheet.conditional_format('{}7:{}{}'.format(colnum, colnum, rowlim),
                                         {'type': 'date',
                                          'criteria': '=',
                                          'value': tday,
                                          'format': format_date_1})

            worksheet.conditional_format('{}7:{}{}'.format(colnum, colnum, rowlim),
                                         {'type': 'date',
                                          'criteria': '<',
                                          'value': tday,
                                          'format': format_date_2})
        #--------------------------------------------------------------------------------------------------------------
        #================================client and comp rateshop ============================================
        # -----------------------------------------------------------------------------------------------
        elif colname in ls:
            worksheet.conditional_format('{}8:{}{}'.format(colnum, colnum, rowlim),
                                         {'type': 'cell',
                                          'criteria': '>',
                                          'value': 0,
                                          'format': format12})
            worksheet.conditional_format('{}8:{}{}'.format(colnum, colnum, rowlim),
                                         {'type': 'cell',
                                          'criteria': '<',
                                          'value': 0,
                                          'format': format11})
            worksheet.conditional_format('{}8:{}{}'.format(colnum, colnum, rowlim),
                                         {'type': 'cell',
                                          'criteria': '=',
                                          'value': 0,
                                          'format': format13})

            worksheet.set_column('{}2:{}{}'.format(colnum, colnum, rowlim), 2)

       #------------------------------------------------------------------------------------------------------------
        # elif (colname in col_name[name_vals['Lowest Rate']:name_vals['Market Trend']-1]):
        # elif (colname not in ls) and ((colname in (df.columns[lwrate:mktrnd]))):
        elif colname in (df.columns[lwrate:mktrnd]):
            worksheet.conditional_format('{}7:{}{}'.format(colnum,colnum,rowlim),
                                         {'type': 'cell',
                                          'criteria' : '>=',
                                          'value' : 0,
                                         'format': format7})
            worksheet.set_column('{}2:{}{}'.format(colnum,colnum,rowlim),15)


            logging.debug('{} Column Conditional Formatting Done.'.format(colname))



        #================================After market trend ============================================
        # elif (colname in col_name[name_vals['Market Trend']:]):
        elif (colname in (df.columns[mktrnd+1:])):
            worksheet.conditional_format('{}7:{}{}'.format(colnum,colnum,rowlim),
                                         {'type': 'cell',
                                          'criteria' : '>=',
                                          'value' : 0,
                                         'format': format6})

            worksheet.conditional_format('{}7:{}{}'.format(colnum,colnum,rowlim),
                                         {'type': 'cell',
                                          'criteria' : '<',
                                          'value' : 0,
                                         'format': border_form})





#===================================================================================================



        #================================setting border for remaining columns ============================================
        
        else:        
            worksheet.conditional_format('{}7:{}{}'.format(colnum,colnum,rowlim),
                                         {'type': 'cell',
                                          'criteria' : '>=',
                                          'value' : 0,
                                         'format': border_form})
    
            worksheet.conditional_format('{}7:{}{}'.format(colnum,colnum,rowlim),
                                         {'type': 'cell',
                                          'criteria' : '<',
                                          'value' : 0,
                                         'format': border_form})

            
            logging.debug('Column border formatting Done')

    #=============================================A.S.Jun23=== GroupingColumns===========================================================

    row_group_ls = [0,1,2,3,4]
    for i in row_group_ls:
        worksheet.set_row(i,None,None,{'level':1})

    current_date_row = df[df['Date'] == tday.date()].index[0] + 5
    for i in range(6,current_date_row,1):
        worksheet.set_row(i,None,None,{'level':1})



    ## Column Grouping
    group_1 = dict({'OTA Revenue': 'ADR OTB'})

    for key, value in group_1.items():
        colnum_1 = name_ind[key]

        colnum_2 = name_ind[value]
        #worksheet.merge_range('{}6:{}6'.format(colnum_1, colnum_2),"OTB",header_format)
        worksheet.set_column('{}:{}'.format(colnum_1, colnum_2), None, None, {'level': 1})


    group_3 = dict({'Rooms Last 7 Days Pickup': 'ADR Last 7 Days Pickup'})

    for key, value in group_3.items():
        colnum_1 = name_ind[key]

        colnum_2 = name_ind[value]

        worksheet.set_column('{}:{}'.format(colnum_1, colnum_2), None, None, {'level': 2})#2
       # worksheet.merge_range('{}6:{}6'.format(colnum_1, colnum_2), "", header_format)  # 21Nov2022#Pickup



    group_30 = dict({'Rooms Last 7 Days Pickup': 'ADR Last 7 Days Pickup'})
    for key, value in group_30.items():

        colnum_1 = name_ind[key]

        colnum_2 = name_ind[value]
        worksheet.set_column('{}:{}'.format(colnum_1, colnum_2), None, None, {'level': 2})

    if 'Convo' in htlname:#A.S. Aug2023
        for col in ['I:I','G:G']:
            worksheet.set_column(col, None, None, {'hidden': 1})

 #  ===================================================================================================================

    # mktrnd=df.columns.get_loc('Market Trend')
    # rt1=mktrnd+2
    # for col_num, value in enumerate(df.columns.values):
    #     if col_num in ((ls) or (rt1)):
    #         worksheet.write(5, col_num, value, header_format_blk_l)
    #     else:
    #         worksheet.write(5, col_num, value, header_format)


    for col_num, value in enumerate(df.columns.values):
        worksheet.write(5, col_num, value, header_format)

    # ----------------Assigning Header Format-----------------------------------

    # ls[:] = [number - 1 for number in ls]
    # for col_num, value in enumerate(df.columns.values):
    #     if col_num in ls:
    #         worksheet.write(5, col_num, value, header_format_blk_l)
    #     else:
    #         worksheet.write(5, col_num, value, header_format)  # adds +1 row
    # ---------------------------------------------------------------------------------------



    
    #===================================== alignment =====================================================================
    alignment = workbook.add_format({ 'align': 'center'})  
    worksheet.set_column('D:{}'.format(col_index[-1]), None, alignment)
    logging.debug('set alignment (center) format to the worksheet Column D')

    #=========================== Market Trend Arrows ==========================================================================
    logging.debug('setting Market Trend Arrows ..')
    mtindx = name_ind['Market Trend']
    # mtindx=df.columns.get_loc('Market Trend')
    logging.debug('Market Trend column index : {}'.format(mtindx)) 
    # worksheet.conditional_format('{}7:{}{}'.format(mtindx,mtindx,rowlim),
    worksheet.conditional_format('{}7:{}{}'.format(mtindx,mtindx,rowlim),
                                 {'type':'icon_set',
                                  'icon_style': '3_arrows',
                                  'icons_only' : True,
                                  'icons': [{'criteria': '>=', 'type': 'number','value': 0},
                                            {'criteria': '<=', 'type': 'number', 'value': 0}]})
    
    worksheet.conditional_format('{}7:{}{}'.format(mtindx,mtindx,rowlim),
                                         {'type': 'cell',
                                          'criteria' : '<',
                                          'value' : 0,
                                         'format': border_form})
    
    logging.debug("set 'icon_style': '3_arrows', Market Trend formatting Done")
    #===========================================================================================================================
    logging.debug("Set A column width to 12")
    worksheet.set_column('{}2:{}{}'.format('A','A',rowlim),12)
     
    worksheet.hide_gridlines(2)
    worksheet.freeze_panes(6, 2)
    logging.debug("Frames Freezed")
    
    worksheet.insert_image('A2', defaultpath+r'\masters\logo\isell.jpg',{'x_offset': 15, 'y_offset': 0,'x_scale': 0.65, 'y_scale': 0.65,'object_position':1})
    logging.debug("isell image inserted")
    
    if htlname in ['Getfam Hotel Addis Ababa', 'Best Western Plus Westlands', 'Hakuna Majiwe Beach Lodge', 'Moja Tuu Luxury Villas & Nature Retreat']:
        worksheet.insert_image('G2', defaultpath+r'\masters\logo\Aleph.png',{'x_offset': 15, 'y_offset': 0,'x_scale': 0.85, 'y_scale': 0.85,'object_position':1})
        logging.debug("getfam logo inserted")
    elif htlname in ['The Boutique Hotel Museum','Mark Hotel Belgrade']:
        worksheet.insert_image('G2', defaultpath+r'\masters\logo\Hotelwise.png',{'x_offset': 15, 'y_offset': 0,'x_scale': 0.85, 'y_scale': 0.85,'object_position':1})
        logging.debug("getfam logo inserted")
    else:
        worksheet.insert_image('G2', defaultpath+r'\masters\logo\logo.jpg',{'x_offset': 15, 'y_offset': 0,'x_scale': 0.85, 'y_scale': 0.85,'object_position':1})
        logging.debug("isell logo inserted")
    
    #=============================================Glossary======================================
    logging.debug("Formatting Glossary ...")
    glossary.to_excel(writer, sheet_name='Glossary',index=False)
    worksheet2 = writer.sheets['Glossary']
    logging.debug("Added Glossary sheet to writer object")
    
    worksheet2.set_column('{}1:{}{}'.format('A','A',rowlim),5)
    worksheet2.set_column('{}1:{}{}'.format('B','B',rowlim),32)
    worksheet2.set_column('{}1:{}{}'.format('C','C',rowlim),46)  
    logging.debug("Set column widths for A,B,C as 5,32 and 46 respectively")
    
    
    gformat = workbook.add_format({ 'border':1})
    gl_shape=glossary.shape
    gl_rows = gl_shape[0]
    
    #assigning conditional formatting to glossary
    worksheet2.conditional_format('{}1:{}{}'.format('A','C',gl_rows+1),
                                         {'type': 'cell',
                                          'criteria' : '>',
                                          'value' : 0,                                
                                         'format': gformat})
    logging.debug("set border to the Glossary")
    
    # Adding Formatting to Glossary
    # Add a bold format to use to highlight cells.
    boldcell = workbook.add_format({'bold': True,'bg_color': '#999999','align': 'center'})
    logging.debug("boldcell format created to highlight and bold the header cells")
    
    #---------------------------Header Formatting------------------------------
    worksheet2.write('A1', 'Sr No', boldcell)    
    worksheet2.write('B1', 'Column Name', boldcell)
    worksheet2.write('C1', 'Column Description', boldcell)       
    logging.debug("Set format of Headers A1,B1,C1 to boldcell")
    
    worksheet2.write('A17', 'Sr No', boldcell)    
    worksheet2.write('B17', 'Pricing Abbreviations', boldcell)
    worksheet2.write('C17', 'Rate Description', boldcell)  
    logging.debug("Set format of Headers A17,B17,C17 to boldcell")
    
    
    worksheet2.write('A45', 'Sr No', boldcell)    
    worksheet2.write('B45', 'Pricing Abbreviations', boldcell)
    worksheet2.write('C45', 'Rate Description', boldcell)  
    logging.debug("Set format of Headers A45,B45,C45 to boldcell") 
      
    #===================================================================================================
    #-----------------------------Grid Addition and Adoption addition---------------------------------------------------
    #===================================================================================================
    if iselltype == 'internal':
        pgdf.to_excel(writer, sheet_name='Grids',index=False) 
        #======================================Add Adoption================================================
        finaladop.to_excel(writer,index=False,sheet_name='iSell_{}'.format(iseldt),startrow=1, startcol=15)
        logging.debug("Grid sheet and adoption sheet added to  for internal iSell")
    
    else:
        pass   
    
    #======================================================================================================
    # RATE SHOP SHEET CREATION
    if iselltype == 'internal':
        rateshopfile.to_excel(writer, sheet_name='RateShop', index=False)
    elif iselltype == 'client':
        rateshopfile.to_excel(writer, sheet_name='RateShop', index=False)
    else:
        pass
    
    logging.debug("RateShop sheet added in internal and client iSells")
    #=====================================================================================================    
    # cm_room_type_df.to_excel(writer, sheet_name='Room Type', startrow=0, startcol=0, index=False)



    writer.save()
    workbook.close()
    logging.debug("iSell Beautification done, workbook saved and closed")
    
def isellbeautify(defaultpath, df, htlname, pth2, name_win2, isellrange, glossary, ftr, pgdf, finaladop,
                  acc_man, rateshopfile, hCap,cm_room_type_df,tday,rng,ls):
    hcap2 = int(hCap*10/100)     
    logging.debug("10% of hotel capacity calculated for highlighting Hotel Availability less than 10%")
    
    try:
        os.chdir(pth2)
        os.mkdir(foldname)
    except:
        pass
    
    pth_1 = pth2 +'\\'+foldname #iSell/tdayfold
    
    try:
        os.chdir(pth_1)
        os.mkdir(acc_man)     #iSell/tdayfold/accman
    except FileExistsError:
        pass
    
    #-------------final pth-----------------------------
    pth = pth_1 +'\\'+acc_man
    
    #----------------debugging paths-------------------------------
    logging.debug("paths variables ::")
    logging.debug('defaultpath : {}'.format(defaultpath))
    logging.debug('pth2 : {}'.format(pth2))
    logging.debug('pth_1 : {}'.format(pth_1))
    logging.debug('pth : {}'.format(pth))  
    #---------------------------------------------------------------

    df = pd.DataFrame(df) 
    logging.debug("final isell assigned to dataframe 'df'")
    # df['Recommended Rate'] = df['Recommended Rate'].replace({[np.inf,-np.inf],0})
    #
    # df['Recommended Rate'] = df['Recommended Rate'].astype('int')

    df.fillna(value='N/A',inplace=True)
    #df['Event'] = df['Event'].astype(str)
    #df['Event'] = df['Event'].replace({'0':np.nan})


    # df['Recommended Rate'] = df['Recommended Rate'].replace({np.nan:0})
    if htlname in ["The Convo Athens Riviera","The Convo Porto Hotel & Apartment","The Convo Syngrou Apartments"]:
        pass
    else:
        df['Event'] = df['Event'].astype(str)
        df['Event'] = df['Event'].replace({'0': np.nan})
        df['Recommended Rate'] = df['Recommended Rate'].astype(str)

    try:
        df['Recommended Rate'] = df['Recommended Rate'].replace({'N/A':np.nan})
    except:
        pass
    df['Recommended Rate']  = df['Recommended Rate'].astype(float)
    df['Recommended Rate']  = df['Recommended Rate'].round(0)
    df.rename(columns={'OTA_Sold':'OTA Sold','LowestRate':'Lowest Rate'},inplace=True)
    
    logging.debug('isell dataframe (df) before beautify function ::')
    logging.debug(df)    
    
    #-----------------------------beautify function call------------------------------    
    beautify(defaultpath, df,'internal',isellrange+rng+6,htlname,pth,glossary,ftr,pgdf,finaladop,rateshopfile,hcap2,cm_room_type_df,tday,ls)
    #--------------------------------------------------------------------------
    logging.info("Internal iSell Dumped")
    logging.debug("Preparing Client iSell as per DOW, iSell Range and Client iSellWindow condition")
    logging.debug('Todays DOW (dow) : {}'.format(dow))
    logging.debug('iSell Range (isellrange) : {}'.format(isellrange))
    logging.debug('Client iSellWindow (clientwindow(180)) : {}'.format(name_win2))   
    
    
    if str(int(name_win2)) != '0':        
        logging.debug('clientwindow(180) != 0 , so Client iSell will be prepared for 180 days')
        df.drop('Recommended Rate',axis=1,inplace=True)
        
        try:
            df.drop('SeasonalRate_y',axis=1,inplace=True)
        except:
            pass
        logging.debug('dropped Recommended Rate and Seasonal Rate Column')
        
        beautify(defaultpath, df,'client',isellrange+rng+6,htlname,pth,glossary,ftr,pgdf,finaladop,rateshopfile,hcap2,cm_room_type_df,tday,ls)
        logging.info("Client iSell Dumped !!!")
        
    else:
        logging.debug('clientwindow(180) == 0, so Client iSell will be prepared for 180/30 days, as per dow') 
           
        if dow in ['Mon','Fri']:
            logging.debug("today's dow is :{} , it falls in ['Mon','Fri']".format(dow))
            logging.debug("so Client iSell will be prepared for 180 days")
            
            df.drop('Recommended Rate',axis=1,inplace=True)
            try:
                df.drop('SeasonalRate_y',axis=1,inplace=True)
            except:
                pass
            logging.debug('dropped Recommended Rate and Seasonal Rate Column')
            
            beautify(defaultpath,df,'client',isellrange+rng+6,htlname,pth,glossary,ftr,pgdf,finaladop,rateshopfile,hcap2,cm_room_type_df,tday,ls)
            logging.info("Client iSell Dumped !!!")
            
        else:
            #----------------------------30 days isell slicing-------------------------------
            logging.debug("today's dow is :{}".format(dow))
            logging.debug("so Client iSell will be prepared for 30 days")            
            df = df.iloc[:91,:]
            logging.debug('iSell dataframe (df) is Sliced till 30 days ::')
            logging.debug(df)            
            
            df.drop('Recommended Rate',axis=1,inplace=True)
            try:
                df.drop('SeasonalRate_y',axis=1,inplace=True)
            except:
                pass
            logging.debug('dropped Recommended Rate and Seasonal Rate Column')
            
            beautify(defaultpath,df,'client',98,htlname,pth,glossary,ftr,pgdf,finaladop,rateshopfile,hcap2,cm_room_type_df,tday,ls)
            logging.info("Client iSell Dumped !!!")


